"""
Excel report generation with formatted Metrics tab and Filtered Data tab.
"""
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter


def _calculate_metrics(df, date_column="date_completed", price_column="price"):
    """Calculate overall, by-quarter, by-month, by-segment metrics."""
    df = df.copy()
    df[date_column] = pd.to_datetime(df[date_column])
    if df[date_column].dt.tz is not None:
        df[date_column] = df[date_column].dt.tz_localize(None)

    overall_avg = df[price_column].mean()
    overall_count = len(df)
    overall_sum = df[price_column].sum()

    # By quarter
    df["quarter"] = df[date_column].dt.to_period("Q")
    avg_by_quarter = df.groupby("quarter")[price_column].agg(["mean", "count", "sum"]).reset_index()
    avg_by_quarter["quarter"] = avg_by_quarter["quarter"].astype(str)
    avg_by_quarter.columns = ["Quarter", "Average Price", "Count", "Total"]

    # By month
    df["_month_sort"] = df[date_column].dt.to_period("M")
    df["month"] = df[date_column].dt.strftime("%B")
    avg_by_month = (
        df.groupby(["_month_sort", "month"])[price_column]
        .agg(["mean", "count", "sum"])
        .reset_index()
    )
    avg_by_month = avg_by_month.sort_values("_month_sort").drop(columns=["_month_sort"])
    avg_by_month.columns = ["Month", "Average Price", "Count", "Total"]

    # By segment
    avg_by_segment = df.groupby("segment")[price_column].agg(["mean", "count", "sum"]).reset_index()
    avg_by_segment.columns = ["Segment", "Average Price", "Count", "Total"]

    return {
        "overall_avg": overall_avg,
        "overall_count": overall_count,
        "overall_sum": overall_sum,
        "by_quarter": avg_by_quarter,
        "by_month": avg_by_month,
        "by_segment": avg_by_segment,
    }


def _build_metrics_sheet_data(metrics, sku_label=None):
    """Build the flat list-of-lists for the Metrics tab."""
    title = f"OVERALL METRICS — {sku_label}" if sku_label else "OVERALL METRICS"
    data = [
        [title, "", "", ""],
        ["Average Price", metrics["overall_avg"], "", ""],
        ["Total Revenue", metrics["overall_sum"], "", ""],
        ["Item Count", metrics["overall_count"], "", ""],
        ["", "", "", ""],
        ["AVERAGE BY QUARTER", "", "", ""],
    ]
    for _, row in metrics["by_quarter"].iterrows():
        data.append([row["Quarter"], row["Average Price"], row["Count"], row["Total"]])
    data.extend([["", "", "", ""], ["AVERAGE BY MONTH", "", "", ""]])
    for _, row in metrics["by_month"].iterrows():
        data.append([row["Month"], row["Average Price"], row["Count"], row["Total"]])
    data.extend([["", "", "", ""], ["AVERAGE BY SEGMENT", "", "", ""]])
    for _, row in metrics["by_segment"].iterrows():
        data.append([row["Segment"], row["Average Price"], row["Count"], row["Total"]])

    return pd.DataFrame(data, columns=["Category", "Average Price", "Count", "Total"])


def _prepare_export_dataframe(df):
    """Strip timezones and add quarter/month/year columns for the data tab."""
    df = df.copy()
    for col in df.columns:
        if isinstance(df[col].dtype, pd.DatetimeTZDtype):
            df[col] = df[col].dt.tz_localize(None)
        elif pd.api.types.is_datetime64_any_dtype(df[col]):
            if df[col].dt.tz is not None:
                df[col] = df[col].dt.tz_localize(None)

    df["date_completed"] = pd.to_datetime(df["date_completed"])
    df["quarter"] = df["date_completed"].dt.to_period("Q").astype(str)
    df["month"] = df["date_completed"].dt.strftime("%B")
    df["year"] = df["date_completed"].dt.year
    return df


def _format_metrics_sheet(ws):
    """Apply formatting to the Metrics worksheet."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18

    section_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    section_font = Font(bold=True, size=12)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if any(
                    kw in cell.value
                    for kw in ["METRICS", "QUARTER", "MONTH", "SEGMENT"]
                ):
                    cell.font = section_font
                    cell.fill = section_fill
                    cell.alignment = Alignment(horizontal="left")
                    # Apply section fill across all columns in the row
                    for sibling in ws[cell.row]:
                        sibling.fill = section_fill

    # Currency formatting on price/total columns (B and D), but not count rows
    count_labels = {"item count", "count"}
    for row_idx in range(2, ws.max_row + 1):
        row_label = str(ws.cell(row_idx, 1).value or "").strip().lower()
        if row_label in count_labels:
            continue
        for col_idx in [2, 4]:  # B=Average Price, D=Total
            cell = ws.cell(row_idx, col_idx)
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE


def _format_data_sheet(ws):
    """Apply formatting to the Filtered Data worksheet."""
    # Auto-adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Header styling
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    dark_blue_border = Border(
        left=Side(color="081F43", style="medium"),
        right=Side(color="081F43", style="medium"),
        top=Side(color="081F43", style="medium"),
        bottom=Side(color="081F43", style="medium"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = dark_blue_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data row borders
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.value:
                cell.border = dark_blue_border

    ws.freeze_panes = "A2"


def generate_report_excel(df, output_path, sku_label=None):
    """Generate the formatted Excel report with Metrics and Filtered Data tabs.

    Args:
        df: DataFrame with line item data (must have 'price', 'date_completed', 'segment')
        output_path: Path for the output .xlsx file
        sku_label: Optional label for the metrics title (e.g., "001-adgm-incorp-001")

    Returns:
        The output_path
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    metrics = _calculate_metrics(df)
    metrics_df = _build_metrics_sheet_data(metrics, sku_label=sku_label)
    export_df = _prepare_export_dataframe(df)

    # Write raw data
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        metrics_df.to_excel(writer, sheet_name="Metrics", index=False)
        export_df.to_excel(writer, sheet_name="Filtered Data", index=False)

    # Apply formatting
    wb = load_workbook(output_path)
    _format_metrics_sheet(wb["Metrics"])
    _format_data_sheet(wb["Filtered Data"])
    wb.save(output_path)

    return output_path
